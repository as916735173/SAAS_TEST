import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def generate_excel_from_case():
    # 配置项（调整路径和文件名格式）
    VERSION = "v2.1.6.7"  # 当前版本号（根据实际项目修改）
    CASE_FILE = os.path.join("common_resources", "data", "case")  # 测试用例JSON文件路径（用户提供的文件）
    EXCEL_OUTPUT_DIR = "test_cases/api"  # Excel输出目录（用户要求的test_cases/api）
    BASE_NAME = f"测试用例_{VERSION}_"  # 文件名前缀（用户要求的格式）

    # 步骤1：检查测试用例JSON文件是否存在
    if not os.path.exists(CASE_FILE):
        print(f"错误：未找到测试用例文件 {os.path.abspath(CASE_FILE)}")
        return

    # 步骤2：读取JSON用例数据
    try:
        with open(CASE_FILE, "r", encoding="utf-8") as f:
            test_cases = json.load(f)
    except json.JSONDecodeError:
        print(f"错误：文件 {CASE_FILE} 非有效JSON格式")
        return
    except Exception as e:
        print(f"错误：读取用例文件失败 - {str(e)}")
        return

    if not test_cases:
        print("错误：测试用例文件内容为空")
        return

    # 步骤3：生成带版本号和序号的Excel文件名
    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)
    existing_excel_files = [f for f in os.listdir(EXCEL_OUTPUT_DIR) 
                            if f.startswith(BASE_NAME) and f.endswith(".xlsx")]
    
    max_num = 0
    for file in existing_excel_files:
        # 正则提取序号（匹配"测试用例_v2.1.6.7_3.xlsx"中的3）
        match = re.search(rf"{BASE_NAME}(\d+)\.xlsx", file)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
    new_num = max_num + 1 if existing_excel_files else 1
    excel_output_file = os.path.join(EXCEL_OUTPUT_DIR, f"{BASE_NAME}{new_num}.xlsx")

    # 步骤4：生成Excel（保持原有表格生成逻辑）
    wb = Workbook()
    ws = wb.active
    ws.title = "车辆管理接口测试用例"

    headers = list(test_cases[0].keys())
    ws.append(headers)

    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment

    for case in test_cases:
        ws.append([case[key] for key in headers])

    # 自动调整列宽
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                continue
        adjusted_width = max(max_length + 2, 15)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(excel_output_file)
    print(f"Excel测试用例文件已成功生成至：{os.path.abspath(excel_output_file)}")

if __name__ == "__main__":
    generate_excel_from_case()